from openpyxl import load_workbook
import re

# 定义一个函数用于校验身份证号码的有效性
def is_valid_idcard(idcard):
    # 验证长度和格式
    if len(idcard) != 18 or not re.match(r"^\d{17}[\dX]$", idcard):
        return False

    # 校验码计算
    nums = [int(x) for x in idcard[:-1]]  # 将身份证号码的前17位转换为整数列表
    weights = (7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2)  # 加权因子
    checksum = sum(a * b for a, b in zip(weights, nums))  # 计算加权和
    charset = '10X98765432'  # 校验码字符集
    return charset[checksum % 11] == idcard[-1].upper()  # 校验码验证

# 加载Excel工作簿并选择第一个工作表
wb = load_workbook('test.xlsx')  # 读取名为'test.xlsx'的Excel文件
sheet = wb['Sheet1']  # 选择工作簿中的第一个工作表'Sheet1'

correct_idcards_col1 = []  # 创建一个空列表用于存储第一列（A列）中正确的身份证号码
correct_idcards_col4 = []  # 创建一个空列表用于存储第四列（D列）中正确的身份证号码

# 从工作表中提取A1-A51和D1-D51列的数据并验证身份证号码
for i in range(2, 52):  # 遍历每列的行，从第2行到第51行
    idcard_col1 = sheet.cell(i, 1).value  # 提取A列（第1列）的数据
    idcard_col4 = sheet.cell(i, 4).value  # 提取D列（第4列）的数据
    
    # 验证A列的身份证号码
    if idcard_col1:
        # 验证前6位是否在指定范围内
        if re.match(r"^(131123|130522)", idcard_col1):
            # 提取出生日期并验证范围
            try:
                year = int(idcard_col1[6:10])
                month = int(idcard_col1[10:12])
                day = int(idcard_col1[12:14])
            except ValueError:
                continue
            
            if (1930 <= year <= 2024) and (1 <= month <= 12) and (1 <= day <= 31) and is_valid_idcard(idcard_col1):
                correct_idcards_col1.append(idcard_col1)  # 将正确的身份证号码添加到列表中
    
    # 验证D列的身份证号码
    if idcard_col4:
        # 验证前6位是否在指定范围内
        if re.match(r"^(131123|130522)", idcard_col4):
            # 提取出生日期并验证范围
            try:
                year = int(idcard_col4[6:10])
                month = int(idcard_col4[10:12])
                day = int(idcard_col4[12:14])
            except ValueError:
                continue
            
            if (1930 <= year <= 2024) and (1 <= month <= 12) and (1 <= day <= 31) and is_valid_idcard(idcard_col4):
                correct_idcards_col4.append(idcard_col4)  # 将正确的身份证号码添加到列表中

# 统计正确的身份证号码数量
correct_count_col1 = len(correct_idcards_col1)
correct_count_col4 = len(correct_idcards_col4)

print("第一列（A列）中正确的身份证号码数量：", correct_count_col1)
print("第一列（A列）中正确的身份证号码列表：", correct_idcards_col1)

print("第四列（D列）中正确的身份证号码数量：", correct_count_col4)
print("第四列（D列）中正确的身份证号码列表：", correct_idcards_col4)
