from openpyxl import load_workbook
import re

# 定义一个函数用于使用Luhn算法校验银行卡号的有效性
def luhn_checksum(card_number):
    def digits_of(n):
        return [int(d) for d in str(n)]
    digits = digits_of(card_number)
    odd_digits = digits[-1::-2]
    even_digits = digits[-2::-2]
    checksum = sum(odd_digits)
    for d in even_digits:
        checksum += sum(digits_of(d * 2))
    return checksum % 10

def is_valid_bankcard(bankcard):
    # 验证长度和格式
    if len(bankcard) != 19 or not bankcard.isdigit():
        return False

    # 验证BIN码是否在指定范围内
    if not re.match(r"^(622481|622513|622250)", bankcard):
        return False

    # 使用Luhn算法验证校验位
    return luhn_checksum(bankcard) == 0

# 加载Excel工作簿并选择第一个工作表
wb = load_workbook('test.xlsx')  # 读取名为'test.xlsx'的Excel文件
sheet = wb['Sheet1']  # 选择工作簿中的第一个工作表'Sheet1'

correct_bankcards_col3 = []  # 创建一个空列表用于存储第三列（C列）中正确的银行卡号
correct_bankcards_col8 = []  # 创建一个空列表用于存储第八列（H列）中正确的银行卡号

# 从工作表中提取C1-C51和H1-H51列的数据并验证银行卡号
for i in range(2, 52):  # 遍历每列的行，从第2行到第51行
    bankcard_col3 = sheet.cell(i, 3).value  # 提取C列（第3列）的数据
    bankcard_col8 = sheet.cell(i, 8).value  # 提取H列（第8列）的数据
    
    # 验证C列的银行卡号
    if bankcard_col3 and is_valid_bankcard(bankcard_col3):
        correct_bankcards_col3.append(bankcard_col3)  # 将正确的银行卡号添加到列表中
    
    # 验证H列的银行卡号
    if bankcard_col8 and is_valid_bankcard(bankcard_col8):
        correct_bankcards_col8.append(bankcard_col8)  # 将正确的银行卡号添加到列表中

# 统计正确的银行卡号数量
correct_count_col3 = len(correct_bankcards_col3)
correct_count_col8 = len(correct_bankcards_col8)

print("第三列（C列）中正确的银行卡号数量：", correct_count_col3)
print("第三列（C列）中正确的银行卡号列表：", correct_bankcards_col3)

print("第八列（H列）中正确的银行卡号数量：", correct_count_col8)
print("第八列（H列）中正确的银行卡号列表：", correct_bankcards_col8)
