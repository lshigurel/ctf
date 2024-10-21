from openpyxl import load_workbook
import re

# 定义一个函数用于校验驾照编号的有效性
def is_valid_drivelicense(drivelicense):
    # 验证长度和格式
    if len(drivelicense) != 12 or not drivelicense.isdigit():
        return False

    # 验证前4位是否在指定范围内
    if not re.match(r"^(1101|1305)", drivelicense):
        return False

    return True

# 加载Excel工作簿并选择第一个工作表
wb = load_workbook('test.xlsx')  # 读取名为'test.xlsx'的Excel文件
sheet = wb['Sheet1']  # 选择工作簿中的第一个工作表'Sheet1'

correct_drivelicenses_col2 = []  # 创建一个空列表用于存储第二列（B列）中正确的驾照编号
correct_drivelicenses_col9 = []  # 创建一个空列表用于存储第九列（I列）中正确的驾照编号

# 从工作表中提取B1-B51和I1-I51列的数据并验证驾照编号
for i in range(2, 52):  # 遍历每列的行，从第2行到第51行
    drivelicense_col2 = sheet.cell(i, 2).value  # 提取B列（第2列）的数据
    drivelicense_col9 = sheet.cell(i, 9).value  # 提取I列（第9列）的数据
    
    # 验证B列的驾照编号
    if drivelicense_col2 and is_valid_drivelicense(drivelicense_col2):
        correct_drivelicenses_col2.append(drivelicense_col2)  # 将正确的驾照编号添加到列表中
    
    # 验证I列的驾照编号
    if drivelicense_col9 and is_valid_drivelicense(drivelicense_col9):
        correct_drivelicenses_col9.append(drivelicense_col9)  # 将正确的驾照编号添加到列表中

# 统计正确的驾照编号数量
correct_count_col2 = len(correct_drivelicenses_col2)
correct_count_col9 = len(correct_drivelicenses_col9)

print("第二列（B列）中正确的驾照编号数量：", correct_count_col2)
print("第二列（B列）中正确的驾照编号列表：", correct_drivelicenses_col2)

print("第九列（I列）中正确的驾照编号数量：", correct_count_col9)
print("第九列（I列）中正确的驾照编号列表：", correct_drivelicenses_col9)
