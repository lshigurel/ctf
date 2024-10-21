from openpyxl import load_workbook
import re

# 定义一个函数用于校验邮箱地址的有效性
def is_valid_email(email):
    # 使用提供的正则表达式
    pattern = r"^[\da-zA-Z][\da-zA-Z]*\w+@(163\.com|qq\.com|hotmail\.com)$"
    return re.match(pattern, email) is not None

# 加载Excel工作簿并选择第一个工作表
wb = load_workbook('test.xlsx')  # 替换为你的文件路径
sheet = wb['Sheet1']  # 选择工作簿中的第一个工作表'Sheet1'

correct_emails_col6 = []  # 创建一个空列表用于存储第六列（F列）中正确的邮箱地址
correct_emails_col7 = []  # 创建一个空列表用于存储第七列（G列）中正确的邮箱地址

# 从工作表中提取F1-F51和G1-G51列的数据并验证邮箱地址
for i in range(2, 52):  # 遍历每列的行，从第2行到第51行
    email_col6 = sheet.cell(i, 6).value  # 提取F列（第6列）的数据
    email_col7 = sheet.cell(i, 7).value  # 提取G列（第7列）的数据
    
    # 验证F列的邮箱地址
    if email_col6 and is_valid_email(str(email_col6).strip()):
        correct_emails_col6.append(email_col6.strip())
    
    # 验证G列的邮箱地址
    if email_col7 and is_valid_email(str(email_col7).strip()):
        correct_emails_col7.append(email_col7.strip())

# 统计正确的邮箱地址数量
correct_count_col6 = len(correct_emails_col6)
correct_count_col7 = len(correct_emails_col7)

print("第六列（F列）中正确的邮箱地址数量：", correct_count_col6)
print("第六列（F列）中正确的邮箱地址列表：", correct_emails_col6)

print("第七列（G列）中正确的邮箱地址数量：", correct_count_col7)
print("第七列（G列）中正确的邮箱地址列表：", correct_emails_col7)