from openpyxl import load_workbook
import re

# 定义一个函数用于校验MAC地址的有效性
def is_valid_mac(mac):
    # 验证长度和格式
    if not re.match(r"^([0-9A-F]{2}:){5}[0-9A-F]{2}$", mac):
        return False

    # 验证前6个十六进制数对是否在指定范围内
    if not re.match(r"^(AA:AD:7C|00:2D:4B|00:1A:5F)", mac):
        return False

    return True

# 加载Excel工作簿并选择第一个工作表
wb = load_workbook('test.xlsx')  # 读取名为'test.xlsx'的Excel文件
sheet = wb['Sheet1']  # 选择工作簿中的第一个工作表'Sheet1'

correct_macs_col5 = []  # 创建一个空列表用于存储第五列（E列）中正确的MAC地址
correct_macs_col10 = []  # 创建一个空列表用于存储第十列（J列）中正确的MAC地址

# 从工作表中提取E1-E51和J1-J51列的数据并验证MAC地址
for i in range(2, 52):  # 遍历每列的行，从第2行到第51行
    mac_col5 = sheet.cell(i, 5).value  # 提取E列（第5列）的数据
    mac_col10 = sheet.cell(i, 10).value  # 提取J列（第10列）的数据
    
    # 验证E列的MAC地址
    if mac_col5 and is_valid_mac(mac_col5):
        correct_macs_col5.append(mac_col5)  # 将正确的MAC地址添加到列表中
    
    # 验证J列的MAC地址
    if mac_col10 and is_valid_mac(mac_col10):
        correct_macs_col10.append(mac_col10)  # 将正确的MAC地址添加到列表中

# 统计正确的MAC地址数量
correct_count_col5 = len(correct_macs_col5)
correct_count_col10 = len(correct_macs_col10)

print("第五列（E列）中正确的MAC地址数量：", correct_count_col5)
print("第五列（E列）中正确的MAC地址列表：", correct_macs_col5)

print("第十列（J列）中正确的MAC地址数量：", correct_count_col10)
print("第十列（J列）中正确的MAC地址列表：", correct_macs_col10)
